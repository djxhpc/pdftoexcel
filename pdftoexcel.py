import os
import fitz  # PyMuPDF
import pdfplumber
import numpy as np
from PIL import Image
from rapidocr_onnxruntime import RapidOCR
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from opencc import OpenCC

def get_native_words(pdfplumber_page):
    """提取頁面中所有的原生數位文字及其精準座標"""
    words = pdfplumber_page.extract_words()
    # 將 pdfplumber 的座標格式標準化，方便後面比對
    # pdfplumber 座標是相對於頁面點數 (points)，我們需要記錄相對比例
    return words

def find_best_native_text(ocr_item, native_words, page_width, page_height, ocr_width, ocr_height):
    """
    智慧比對：拿 OCR 的座標去原生文字庫裡撈字。
    如果重疊度高，就直接用 100% 正確的原生文字取代 OCR 辨識結果。
    """
    ocr_x_min, ocr_x_max = ocr_item['x_min'], ocr_item['x_max']
    ocr_y_min, ocr_y_max = ocr_item['y_min'], ocr_item['y_max']
    
    # 轉換 OCR 座標系到 pdfplumber 座標系
    scale_x = page_width / ocr_width
    scale_y = page_height / ocr_height
    
    matched_texts = []
    for nw in native_words:
        nw_x_min = nw['x0']
        nw_x_max = nw['x1']
        nw_y_min = nw['top']
        nw_y_max = nw['bottom']
        
        # 檢查視覺邊框是否有交集（重疊）
        match_x = max(ocr_x_min * scale_x, nw_x_min) < min(ocr_x_max * scale_x, nw_x_max)
        match_y = max(ocr_y_min * scale_y, nw_y_min) < min(ocr_y_max * scale_y, nw_y_max)
        
        if match_x and match_y:
            matched_texts.append(nw)
            
    if matched_texts:
        # 依照從左到右排序組合成完整的字串
        matched_texts.sort(key=lambda x: x['x0'])
        return "".join([w['text'] for w in matched_texts])
        
    return None # 沒找到原生文字，代表這真的是純圖片

def parse_pdf_table_ultimate(pdf_path, output_excel_path, y_threshold=20, x_threshold=25):
    """
    終極版：表格結構完全依賴視覺（不漏圖），但內容智慧融合原生文字（不寫錯字）
    """
    engine = RapidOCR()
    cc = OpenCC('s2tw')
    wb = Workbook()
    wb.remove(wb.active)
    
    doc_fitz = fitz.open(pdf_path)
    
    with pdfplumber.open(pdf_path) as doc_plumber:
        for page_num in range(len(doc_fitz)):
            print(f"正在全景掃描第 {page_num + 1} / {len(doc_fitz)} 頁...")
            page_fitz = doc_fitz[page_num]
            page_plumber = doc_plumber.pages[page_num]
            
            # 1. 取得原生數位文字庫
            native_words = get_native_words(page_plumber)
            
            # 2. 將頁面轉成高解析度圖片，讓 OCR 抓出「所有看得見的表格與文字結構」
            pix = page_fitz.get_pixmap(matrix=fitz.Matrix(3, 3))
            img_data = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img_np = np.array(img_data)
            ocr_height, ocr_width = img_np.shape[0], img_np.shape[1]
            
            result, _ = engine(img_np)
            if not result:
                print(f"第 {page_num + 1} 頁空白或無法辨識。")
                continue
                
            # 3. 解析 OCR 結果並與原生文字融合
            raw_data = []
            for item in result:
                box, text, _ = item
                x_min = min(pt[0] for pt in box)
                x_max = max(pt[0] for pt in box)
                y_min = min(pt[1] for pt in box)
                y_max = max(pt[1] for pt in box)
                
                ocr_item = {'x_min': x_min, 'x_max': x_max, 'y_min': y_min, 'y_max': y_max}
                
                # 智慧撈取原生文字
                native_text = find_best_native_text(
                    ocr_item, native_words, 
                    page_plumber.width, page_plumber.height, 
                    ocr_width, ocr_height
                )
                
                # 如果有原生文字就用原生的（100%準確），沒有就用 OCR 辨識出來的
                final_text = native_text if native_text else cc.convert(text)
                
                raw_data.append({
                    'x_min': x_min, 'x_max': x_max,
                    'y_min': y_min, 'y_max': y_max,
                    'text': final_text
                })

            # 4. 根據 Y 座標進行多行/多列的還原演算法
            raw_data.sort(key=lambda item: item['y_min'])
            initial_rows = []
            current_row = []
            last_y = -1
            
            for item in raw_data:
                if last_y == -1 or abs(item['y_min'] - last_y) <= y_threshold:
                    current_row.append(item)
                else:
                    initial_rows.append(current_row)
                    current_row = [item]
                last_y = item['y_min']
            if current_row:
                initial_rows.append(current_row)

            # 合併同儲存格換行的文字
            optimized_rows = []
            skip_next = False
            for idx in range(len(initial_rows)):
                if skip_next:
                    skip_next = False
                    continue
                current_r = initial_rows[idx]
                if idx + 1 < len(initial_rows):
                    next_r = initial_rows[idx + 1]
                    max_c_ymax = max(item['y_max'] for item in current_r)
                    min_n_ymin = min(item['y_min'] for item in next_r)
                    
                    if 0 <= (min_n_ymin - max_c_ymax) <= y_threshold * 1.5:
                        merged_row = []
                        next_used = set()
                        for c_item in current_r:
                            for n_idx, n_item in enumerate(next_r):
                                overlap = min(c_item['x_max'], n_item['x_max']) - max(c_item['x_min'], n_item['x_min'])
                                if overlap > 5:
                                    c_item['text'] = c_item['text'] + "\n" + n_item['text']
                                    c_item['y_max'] = n_item['y_max']
                                    next_used.add(n_idx)
                                    break
                            merged_row.append(c_item)
                        for n_idx, n_item in enumerate(next_r):
                            if n_idx not in next_used:
                                merged_row.append(n_item)
                        current_r = merged_row
                        skip_next = True
                optimized_rows.append(current_r)

            # 5. 寫入 Excel 並保持比例
            ws = wb.create_sheet(title=f"Page_{page_num + 1}")
            ws.views.sheetView[0].showGridLines = True
            
            for row_idx, row_data in enumerate(optimized_rows, start=1):
                row_data.sort(key=lambda item: item['x_min'])
                current_col_idx = 1
                last_x = 0
                for item in row_data:
                    if last_x != 0 and (item['x_min'] - last_x) > x_threshold * 3:
                        skip_cols = int((item['x_min'] - last_x) / (x_threshold * 4))
                        current_col_idx += max(1, skip_cols)
                    
                    cell = ws.cell(row=row_idx, column=current_col_idx, value=item['text'])
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    last_x = item['x_max']
                    current_col_idx += 1

            # 自動調整欄寬
            for col in ws.columns:
                max_len = max(max([len(line) for line in str(cell.value or '').split('\n')]) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len * 1.8, 12)

    wb.save(output_excel_path)
    print(f"\n【終極全景融合版】處理完畢！已儲存至 {output_excel_path}")

if __name__ == "__main__":
    pdf_file = r"D:\work\0601\1150213.pdf" 
    output_excel = r"D:\work\0601\pdf_perfect_output.xlsx"
    
    if os.path.exists(pdf_file):
        parse_pdf_table_ultimate(pdf_file, output_excel)
    else:
        print("找不到檔案，請確認路徑。")
